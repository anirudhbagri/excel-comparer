from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.cell import Cell

def compare_cells(c1: Cell, c2: Cell):
    #TODO: ignore whitespace and add case-insensitive comparison
    if c1.value != c2.value:
        print(f"[w1'{c1.coordinate}' != w2'{c2.coordinate}'] {c1.value} != {c2.value}")

def compare_columns(ws1: Worksheet, ws2: Worksheet, col1: str, col2: str):
    for cell_1, cell_2 in zip(ws1[col1], ws2[col2]):
        compare_cells(cell_1, cell_2)
    if len(ws1[col1]) > len(ws2[col2]):
        print(f"[ws1'{col1}' has more rows than ws2'{col2}', skipping extra rows]\n")
    if len(ws1[col1]) < len(ws2[col2]):
        print(f"[ws2'{col2}' has more rows than ws1'{col1}', skipping extra rows]\n")

def compare_entire_worksheets(ws1: Worksheet, ws2: Worksheet):
    ws1_rows, ws2_rows = ws1.max_row, ws2.max_row
    ws1_cols, ws2_cols = ws1.max_column, ws2.max_column
    rows = min(ws1_rows, ws2_rows)
    cols = min(ws1_cols, ws2_cols)
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            compare_cells(ws1.cell(i, j), ws2.cell(i, j))
    if ws1_rows > ws2_rows:
        print(f"[ws1'{ws1.title}' has more rows than ws2'{ws2.title}', skipping extra rows]")
    if ws1_rows < ws2_rows:
        print(f"[ws2'{ws2.title}' has more rows than ws1'{ws1.title}', skipping extra rows]")
    if ws1_cols > ws2_cols:
        print(f"[ws1'{ws1.title}' has more columns than ws2'{ws2.title}', skipping extra columns]")
    if ws1_cols < ws2_cols:
        print(f"[ws2'{ws2.title}' has more columns than ws1'{ws1.title}', skipping extra columns]")

def compare_workbooks(
    w1: Workbook,
    w2: Workbook,
    s1: str = None,
    s2: str = None,
    cols1: str = None,
    cols2: str = None,
):
    ws1: Worksheet = w1[s1] if s1 else w1.active
    ws2: Worksheet = w2[s2] if s2 else w2.active
    if not cols1 and not cols2:
        compare_entire_worksheets(ws1, ws2)
    elif len(cols1.split(",")) != len(cols2.split(",")):
        print("Columns length don't match, please check the cols input!")
        return
    else:
        for c1, c2 in zip([x.strip() for x in cols1.split(",")], [x.strip() for x in cols2.split(",")]):
            compare_columns(ws1, ws2, c1, c2)

def main(w1, w2, s1=None, s2=None, cols1=None, cols2=None):
    ws1: Workbook = load_workbook(w1)
    ws2: Workbook = load_workbook(w2)
    print("Starting comparison...\n")
    compare_workbooks(ws1, ws2, s1=s1, s2=s2, cols1=cols1, cols2=cols2)
    print("\nComparison finished!")
